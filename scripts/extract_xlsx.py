#!/usr/bin/env python3
"""Extract IPAM data from the source xlsx into a normalized JSON seed.

Run:  python3 scripts/extract_xlsx.py
Output: scripts/seed.json
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "Controle de IP - LAN.xlsx"
OUT = ROOT / "scripts" / "seed.json"

# Per-site sheets contain horizontally-stacked subnets.
# Each subnet block is 4 columns (IP / TIPO / HOSTNAME / FUNÇÃO) followed by 1 blank column.
LAN_SHEETS = [
    "LAN-DUO",
    "LAN-MO",
    "BAGRE-SP3",
    "SIRESP_SP-SP3",
    "CORE_MT-SP3",
    "SALUX_AM-SP3",
    "RACK_DUO-SP3",
    "NOVA LIMA_MG-SP3",
    "MARANHÃO_MA-SP3",
]

# Sheets used for VLAN/range master data.
EQUINIX_SHEET = ""
AZURE_SHEET = "Azure - SCE"
RANGES_SHEET = "Controle - Ranges de IP"
CIDR_SHEET = "Lista de CIDR"


CIDR_RE = re.compile(r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}/\d{1,2})")
IP_RE = re.compile(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$")


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v


def parse_lan_sheet(ws, site_code):
    """A LAN sheet has up to 5 horizontally-stacked subnet blocks.

    - Row 2:  block title (e.g. "CONTROLE DE IP - LAN-DUO-HA")
    - Row 3:  range string (e.g. "RANGE - 10.230.0.0/28 e 10.230.0.17/28")
    - Row 6:  column headers (IP | TIPO | HOSTNAME | FUNÇÃO)
    - Row 7+: data

    Title and range may live in different columns relative to the IP column,
    so we anchor on the "IP" header in row 6 and search nearby for context.
    """
    blocks = []
    max_col = ws.max_column

    def find_nearby(row, ip_col, predicate, span=3):
        """Search columns [ip_col-span .. ip_col+span] for a value matching predicate."""
        for delta in range(-span, span + 1):
            c = ip_col + delta
            if 1 <= c <= max_col:
                v = clean(ws.cell(row=row, column=c).value)
                if v and predicate(v):
                    return v
        return None

    for ip_col in range(1, max_col + 1):
        if clean(ws.cell(row=6, column=ip_col).value) != "IP":
            continue
        title = find_nearby(2, ip_col, lambda v: isinstance(v, str) and "CONTROLE DE IP" in v.upper())
        rng = find_nearby(3, ip_col, lambda v: isinstance(v, str) and "RANGE" in v.upper())
        if not title and not rng:
            continue
        block_name = (title or f"{site_code}-block-{ip_col}").replace("CONTROLE DE IP - ", "").strip()
        cidrs = CIDR_RE.findall(rng) if rng else []
        ips = []
        for r in range(7, ws.max_row + 1):
            ip = clean(ws.cell(row=r, column=ip_col).value)
            tipo = clean(ws.cell(row=r, column=ip_col + 1).value)
            host = clean(ws.cell(row=r, column=ip_col + 2).value)
            func = clean(ws.cell(row=r, column=ip_col + 3).value)
            if not ip and not tipo and not host and not func:
                continue
            if ip and not IP_RE.match(str(ip)):
                continue
            ips.append(
                {
                    "address": str(ip) if ip else None,
                    "type": tipo,
                    "hostname": host,
                    "function": func,
                }
            )
        blocks.append(
            {
                "site_code": site_code,
                "name": block_name,
                "range_label": rng,
                "cidrs": cidrs,
                "ips": ips,
            }
        )
    return blocks


def parse_equinix(ws):
    """ has VLAN definitions starting at row 6.
    Header row 5: Name | Vlan ID | Network Address | Usage Host Range | Broadcast.
    """
    out = []
    for r in range(6, ws.max_row + 1):
        name = clean(ws.cell(row=r, column=2).value)
        vlan = clean(ws.cell(row=r, column=3).value)
        net = clean(ws.cell(row=r, column=4).value)
        usage = clean(ws.cell(row=r, column=5).value)
        bcast = clean(ws.cell(row=r, column=6).value)
        if not name and not net:
            continue
        out.append(
            {
                "name": name,
                "vlan_id": vlan,
                "network": net,
                "usage": usage,
                "broadcast": bcast,
            }
        )
    return out


def parse_azure_subnets(ws):
    """Azure - SCE has Vnet rows in cols 2-5 starting row 6."""
    out = []
    for r in range(6, ws.max_row + 1):
        name = clean(ws.cell(row=r, column=2).value)
        net = clean(ws.cell(row=r, column=3).value)
        usage = clean(ws.cell(row=r, column=4).value)
        bcast = clean(ws.cell(row=r, column=5).value)
        if not name and not net:
            continue
        out.append({"name": name, "network": net, "usage": usage, "broadcast": bcast})
    return out


def parse_azure_firewall(ws):
    """Firewall rules live in cols 9-17 of the Azure sheet."""
    rules = []
    # find header row (contains 'Sentido')
    header_row = None
    for r in range(1, ws.max_row + 1):
        if clean(ws.cell(row=r, column=9).value) == "Sentido":
            header_row = r
            break
    if header_row is None:
        return rules
    for r in range(header_row + 1, ws.max_row + 1):
        sentido = clean(ws.cell(row=r, column=9).value)
        in_iface = clean(ws.cell(row=r, column=10).value)
        out_iface = clean(ws.cell(row=r, column=11).value)
        src = clean(ws.cell(row=r, column=12).value)
        dst = clean(ws.cell(row=r, column=13).value)
        port = clean(ws.cell(row=r, column=14).value)
        service = clean(ws.cell(row=r, column=16).value)
        proto = clean(ws.cell(row=r, column=17).value)
        if not any([sentido, src, dst, service]):
            continue
        rules.append(
            {
                "direction": sentido,
                "in_iface": in_iface,
                "out_iface": out_iface,
                "src": src,
                "dst": dst,
                "port": port,
                "service": service,
                "protocol": proto,
            }
        )
    return rules


def parse_master_ranges(ws):
    """Best-effort extraction of master range list. The sheet is heterogeneous
    so we just emit rows that look like CIDRs."""
    out = []
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = clean(ws.cell(row=r, column=c).value)
            if not v or not isinstance(v, str):
                continue
            m = CIDR_RE.search(v)
            if m:
                cidr = m.group(1)
                # neighbour cell (right) often holds description
                desc = clean(ws.cell(row=r, column=c + 1).value)
                out.append({"cidr": cidr, "description": desc, "row": r, "col": c})
                break  # one cidr per row max
    return out


def parse_cidr_table(ws):
    out = []
    for r in range(3, ws.max_row + 1):
        mask = clean(ws.cell(row=r, column=2).value)
        prefix = clean(ws.cell(row=r, column=3).value)
        total = clean(ws.cell(row=r, column=4).value)
        usable = clean(ws.cell(row=r, column=5).value)
        nets = clean(ws.cell(row=r, column=6).value)
        if not prefix:
            continue
        out.append(
            {
                "mask": mask,
                "prefix": prefix,
                "total": total,
                "usable": usable,
                "networks_per_24": nets,
            }
        )
    return out


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    sites = []
    for sheet in LAN_SHEETS:
        ws = wb[sheet]
        blocks = parse_lan_sheet(ws, sheet)
        sites.append(
            {
                "code": sheet,
                "name": sheet,
                "subnets": blocks,
            }
        )

    seed = {
        "sites": sites,
        "equinix_vlans": parse_equinix(wb[EQUINIX_SHEET]),
        "azure_subnets": parse_azure_subnets(wb[AZURE_SHEET]),
        "firewall_rules": parse_azure_firewall(wb[AZURE_SHEET]),
        "master_ranges": parse_master_ranges(wb[RANGES_SHEET]),
        "cidr_reference": parse_cidr_table(wb[CIDR_SHEET]),
    }

    OUT.write_text(json.dumps(seed, ensure_ascii=False, indent=2))
    # Stats
    total_ips = sum(len(b["ips"]) for s in sites for b in s["subnets"])
    total_subnets = sum(len(s["subnets"]) for s in sites)
    print(f"Sites: {len(sites)}  Subnets: {total_subnets}  IPs: {total_ips}")
    print(f"Equinix VLANs: {len(seed['equinix_vlans'])}")
    print(f"Azure subnets: {len(seed['azure_subnets'])}")
    print(f"Firewall rules: {len(seed['firewall_rules'])}")
    print(f"Master ranges: {len(seed['master_ranges'])}")
    print(f"CIDR refs: {len(seed['cidr_reference'])}")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
